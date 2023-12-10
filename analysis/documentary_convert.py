

# convert "data/documentary_fills.xlsx" to "data/documentary_fills.json"
import openpyxl
import json

wb = openpyxl.load_workbook("data/documentary_fills.xlsx")
sheet = wb["Sheet1"]
max_row = sheet.max_row

data = []
for row in range(2, max_row + 1):
    paper_id = sheet.cell(row=row, column=1).value
    title = sheet.cell(row=row, column=2).value
    artifact_url = sheet.cell(row=row, column=3).value
    completeness = sheet.cell(row=row, column=4).value
    structure = sheet.cell(row=row, column=5).value
    usability = sheet.cell(row=row, column=6).value
    use_example = sheet.cell(row=row, column=7).value
    certificate = sheet.cell(row=row, column=8).value
    contact = sheet.cell(row=row, column=9).value
    data.append({
        "paper_id": paper_id,
        "title": title,
        "artifact_url": artifact_url,
        "completeness": completeness,
        "structure": structure,
        "usability": usability,
        "use_example": use_example,
        "certificate": certificate,
        "contact": contact
    })
with open("data/documentary_fills.json", "w") as f:
    json.dump(data, f, indent=4)