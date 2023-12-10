'''
open the "data/papers.xlsx" file and convert it into "data/papers.json" and "data/papers.csv"
the "data/papers.xlsx" file have six columns:
- paper_id
- title
- artifact_url
- artifact_url_valid
- URL_location
- URL_format
'''

import openpyxl

wb = openpyxl.load_workbook("data/papers.xlsx")
sheet = wb["Sheet1"]
max_row = sheet.max_row

data = []

for row in range(2, max_row + 1):
    paper_id = sheet.cell(row=row, column=1).value
    title = sheet.cell(row=row, column=2).value
    artifact_url = sheet.cell(row=row, column=3).value
    artifact_url_valid = sheet.cell(row=row, column=4).value
    url_location = sheet.cell(row=row, column=5).value
    url_format = sheet.cell(row=row, column=6).value
    data.append({
        "paper_id": paper_id,
        "title": title,
        "artifact_url": artifact_url if artifact_url_valid else "",
        "artifact_url_valid": artifact_url_valid if artifact_url else "",
        "url_location": url_location if url_location else "",
        "url_format": url_format if url_format else ""
    })

# save to "data/papers.json"
import json
with open("data/papers.json", "w") as f:
    json.dump(data, f, indent=4)

# save to "data/papers.csv"
with open("data/papers.csv", "w") as f:
    f.write("\t".join(data[0].keys()) + "\n")
    for d in data:
        f.write("\t".join(d.values()) + "\n")




