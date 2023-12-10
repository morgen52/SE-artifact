import json
import openpyxl

with open("data/processed_papers.json", "r") as f:
    papers = json.load(f)

# create the documentary.xlsx template for manual annotation
doc_wb = openpyxl.Workbook()
doc_sheet = doc_wb.active
doc_sheet.title = "Sheet1"
doc_sheet.cell(row=1, column=1).value = "paper_id"
doc_sheet.cell(row=1, column=2).value = "title"
doc_sheet.cell(row=1, column=3).value = "artifact_url"
# Completeness,Strucure,Usability,Use example,certificate,contact
doc_sheet.cell(row=1, column=4).value = "Completeness"
doc_sheet.cell(row=1, column=5).value = "Structure"
doc_sheet.cell(row=1, column=6).value = "Usability"
doc_sheet.cell(row=1, column=7).value = "Use example"
doc_sheet.cell(row=1, column=8).value = "Certificate"
doc_sheet.cell(row=1, column=9).value = "Contact"
row = 2
for paper in papers:
    id = paper["paper_id"] # only consider papers with paper_id of ICSE
    repo_name = paper["repo_name"] # only consider papers with repo_name of Github
    if ("ICSE" in id) and repo_name:
        doc_sheet.cell(row=row, column=1).value = paper["paper_id"]
        doc_sheet.cell(row=row, column=2).value = paper["title"]
        doc_sheet.cell(row=row, column=3).value = f"https://github.com/{repo_name}"
        row += 1
doc_wb.save("data/documentary.xlsx")

